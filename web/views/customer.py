import os
import mimetypes

from django.conf import settings
from django.http import FileResponse
from django.shortcuts import render, redirect
from openpyxl.reader.excel import load_workbook

from web import models
from web.forms.customer import CustomerForm


def customer_list(request):
    """
    客户列表
    :return:
    """
    data_list = models.Customer.objects.all()

    return render(request, 'customer_list.html', {'data_list': data_list})


def customer_add(request):
    """
    编辑客户
    :return:
    """
    if request.method == 'GET':
        form = CustomerForm()
        return render(request, 'customer_edit.html', {'form': form})
    form = CustomerForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/customer/list/')
    return render(request, 'customer_edit.html', {'form': form})


def customer_edit(request, cid):
    """
    新增客户
    :return:
    """
    obj = models.Customer.objects.get(id=cid)
    if request.method == 'GET':
        form = CustomerForm(instance=obj)
        return render(request, 'customer_add.html', {'form': form})
    form = CustomerForm(data=request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return redirect('/customer/list/')
    return render(request, 'customer_add.html', {'form': form})


def customer_del(request, cid):
    """
    删除客户
    :param request:
    :param cid:
    :return:
    """
    models.Customer.objects.filter(id=cid).delete()
    return redirect('/customer/list/')


def customer_import(request):
    """
    批量导入
    :param request:
    :return:
    """

    if request.method == 'GET':
        return render(request, 'customer_import.html')

    context = {'status': True, 'msg': '导入成功'}
    try:
        customer_excel = request.FILES.get('customer_excel')
        """
        打开上传的Excel文件，并读取内容
        注：打开本地文件时，可以使用：workbook = xlrd.open_workbook(filename='本地文件路径.xlsx')
        """
        wb=load_workbook(customer_excel)
        sheet=wb.worksheets[0]

        for row in sheet.iter_rows(min_row=2):
            name=row[0].value,
            age=row[1].value,
            email=row[2].value,
            company=row[3].value,
            models.Customer.objects.create(name=name,age=age,email=email,company=company)
    except Exception as e:
        context['status'] = False
        context['msg'] = '导入失败'

    return render(request, 'customer_import.html', context)


def customer_tpl(request):
    """
    下载批量导入Excel列表
    :param request:
    :return:
    """
    tpl_path = os.path.join(settings.BASE_DIR, 'web', 'files', '批量导入客户模板.xlsx')
    content_type = mimetypes.guess_type(tpl_path)[0]
    print(content_type)
    response = FileResponse(open(tpl_path, mode='rb'), content_type=content_type)
    response['Content-Disposition'] = "attachment;filename=%s" % 'customer_excel_tpl.xlsx'
    return response
